import openpyxl
from openpyxl.styles import PatternFill
import random


def generate_random_color():
    return PatternFill(start_color="%06x" % random.randint(0, 0xFFFFFF),
                       end_color="%06x" % random.randint(0, 0xFFFFFF),
                       fill_type='solid')


def apply_fill_colors(input_file_path):
    workbook = openpyxl.load_workbook(input_file_path)
    sheet = workbook.active

    author_colors = {}
    author_counts = {}

    for row in sheet.iter_rows(min_row=2):
        formatted_author_name = row[2].value

        if formatted_author_name in author_counts:
            author_counts[formatted_author_name] += 1
        else:
            author_counts[formatted_author_name] = 1

        if author_counts[formatted_author_name] >= 2:
            if formatted_author_name in author_colors:
                fill_color = author_colors[formatted_author_name]
            else:
                fill_color = generate_random_color()
                author_colors[formatted_author_name] = fill_color

    for row in sheet.iter_rows(min_row=2):
        formatted_author_name = row[2].value
        fill_color = author_colors.get(formatted_author_name)

        if fill_color:
            for cell in row:
                cell.fill = fill_color
    sheet.delete_cols(3)

    workbook.save(input_file_path)
    workbook.close()

if __name__ == "__main__":
    apply_fill_colors('author_filtered_data.xlsx')
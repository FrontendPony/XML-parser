import openpyxl
import pandas as pd
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import random


def generate_random_color():
    return PatternFill(start_color="%06x" % random.randint(0, 0xFFFFFF),
                       end_color="%06x" % random.randint(0, 0xFFFFFF),
                       fill_type='solid')


def add_dropdown_with_ids_to_excel(data_arrays, file_name):
    try:
        try:
            workbook = openpyxl.load_workbook(file_name)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        sheet = workbook.active

        # Add new name to the first row of column F
        sheet[f'G1'].value = 'author_id_choice'
        sheet[f'G1'].font = Font(bold=True)

        row = 2
        for data_array in data_arrays:
            col = 1
            for element in data_array:
                col += 1
                people_ids = [data_array[i] for i in range(1, len(data_array), 4)]
                dv = DataValidation(type="list", formula1=f'"{",".join(map(str, people_ids))}"')
                sheet.add_data_validation(dv)
                dv.add(sheet[f'G{row}'])
                if col > 4:
                    col = 1
                    row += 1

        workbook.save(file_name)

    except Exception as e:
        print(f"An error occurred: {e}")


